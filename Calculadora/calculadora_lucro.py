import pandas as pd
import sys
import os
import warnings
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Ignorar o aviso de estilo padrão do openpyxl
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

def calcular_lucro_semanal(caminho_arquivo):
    print(f"Analisando o arquivo: {caminho_arquivo}\n")
    
    try:
        # Carregar apenas a planilha de transações
        df_trans = pd.read_excel(caminho_arquivo, sheet_name='Transaction Report', skiprows=17)
    except Exception as e:
        print(f"Erro ao ler o arquivo Excel: {e}")
        return

    # Tabela auxiliar interna (fixa no código) com os dados extraídos
    dados_auxiliares = [
        {'Valor final': 15.99, 'Custo': 10.33, 'Lucro': 5.66},
        {'Valor final': 23.52, 'Custo': 14.81, 'Lucro': 8.71},
        {'Valor final': 31.36, 'Custo': 19.29, 'Lucro': 12.07},
        {'Valor final': 39.19, 'Custo': 23.77, 'Lucro': 15.42},
        {'Valor final': 47.99, 'Custo': 28.25, 'Lucro': 19.74},
        {'Valor final': 55.99, 'Custo': 32.73, 'Lucro': 23.26},
        {'Valor final': 59.99, 'Custo': 37.21, 'Lucro': 22.78},
        {'Valor final': 75.98, 'Custo': 47.54, 'Lucro': 28.44},
        {'Valor final': 31.98, 'Custo': 20.66, 'Lucro': 11.32},
        {'Valor final': 119.98, 'Custo': 74.42, 'Lucro': 45.56},
        {'Valor final': 47.04, 'Custo': 29.62, 'Lucro': 17.42},
        {'Valor final': 19.99, 'Custo': 10.33, 'Lucro': 9.66},
        {'Valor final': 39.51, 'Custo': 25.14, 'Lucro': 14.37},
        {'Valor final': 91.36, 'Custo': 56.50, 'Lucro': 34.86},
        {'Valor final': 61.39, 'Custo': 28.25, 'Lucro': 33.14},
        {'Valor final': 94.08, 'Custo': 57.87, 'Lucro': 36.21},
        {'Valor final': 95.98, 'Custo': 56.50, 'Lucro': 39.48},
        {'Valor final': 31.99, 'Custo': 14.81, 'Lucro': 17.18},
        {'Valor final': 43.99, 'Custo': 19.29, 'Lucro': 24.70},
        {'Valor final': 26.80, 'Custo': 14.81, 'Lucro': 11.99},
        {'Valor final': 31.50, 'Custo': 14.81, 'Lucro': 16.69},
        {'Valor final': 17.59, 'Custo': 10.33, 'Lucro': 7.26},
        {'Valor final': 41.59, 'Custo': 23.77, 'Lucro': 17.82},
        {'Valor final': 70.29, 'Custo': 32.73, 'Lucro': 37.56},
        {'Valor final': 34.80, 'Custo': 19.29, 'Lucro': 15.51},
        {'Valor final': 55.50, 'Custo': 23.77, 'Lucro': 31.73},
        {'Valor final': 39.98, 'Custo': 20.66, 'Lucro': 19.32},
        {'Valor final': 43.50, 'Custo': 19.29, 'Lucro': 24.21},
        {'Valor final': 63.49, 'Custo': 29.62, 'Lucro': 33.87},
        {'Valor final': 82.33, 'Custo': 37.21, 'Lucro': 45.12},
        {'Valor final': 19.50, 'Custo': 10.33, 'Lucro': 9.17},
        {'Valor final': 51.19, 'Custo': 28.25, 'Lucro': 22.94},
        {'Valor final': 35.18, 'Custo': 20.66, 'Lucro': 14.52},
        {'Valor final': 80.40, 'Custo': 44.43, 'Lucro': 35.97},
        {'Valor final': 83.16, 'Custo': 37.21, 'Lucro': 45.95},
        {'Valor final': 249.00, 'Custo': 111.63, 'Lucro': 137.37},
        {'Valor final': 33.44, 'Custo': 12.49, 'Lucro': 20.95},
        {'Valor final': 44.39, 'Custo': 25.14, 'Lucro': 19.25},
        {'Valor final': 48.79, 'Custo': 18.05, 'Lucro': 30.74},
        {'Valor final': 86.79, 'Custo': 47.54, 'Lucro': 39.25},
        {'Valor final': 68.78, 'Custo': 28.25, 'Lucro': 40.53},
        {'Valor final': 139.20, 'Custo': 77.16, 'Lucro': 62.04},
        {'Valor final': 77.58, 'Custo': 43.06, 'Lucro': 34.52},
        {'Valor final': 102.38, 'Custo': 56.50, 'Lucro': 45.88},
        {'Valor final': 69.60, 'Custo': 38.58, 'Lucro': 31.02},


        {'Valor final': 14.80, 'Custo': 6.93, 'Lucro': 7.87}, #sacola pequena
        {'Valor final': 22.24, 'Custo': 9.71, 'Lucro': 12.53},
        {'Valor final': 29.44, 'Custo': 12.49, 'Lucro': 16.95},
        {'Valor final': 36.40, 'Custo': 15.27, 'Lucro': 21.13},
        {'Valor final': 43.12, 'Custo': 18.05, 'Lucro': 25.07},
        {'Valor final': 50.24, 'Custo': 20.83, 'Lucro': 29.41},
        {'Valor final': 54.40, 'Custo': 23.61, 'Lucro': 30.79}
    ]
    df_aux = pd.DataFrame(dados_auxiliares)
    df_aux = df_aux.sort_values('Valor final')

    # Limpar e preparar a planilha de Transações
    df_trans.columns = df_trans.columns.str.strip()
    
    if 'Tipo de transação' not in df_trans.columns:
        print("A coluna 'Tipo de transação' não foi encontrada. Verifique o formato do relatório.")
        return

    # Filtrar apenas as rendas de pedidos (entradas)
    df_renda = df_trans[df_trans['Tipo de transação'] == 'Renda do pedido'].copy()
    df_renda['Valor'] = pd.to_numeric(df_renda['Valor'], errors='coerce')
    df_renda = df_renda.dropna(subset=['Valor']).sort_values('Valor')
    
    # Converter a coluna de Data para datetime
    df_renda['Data'] = pd.to_datetime(df_renda['Data'], errors='coerce')

    # Fazer o cruzamento (merge) aproximado para lidar com diferenças de centavos (ex: 31.99 e 31.98)
    merged = pd.merge_asof(
        df_renda, 
        df_aux, 
        left_on='Valor', 
        right_on='Valor final', 
        tolerance=0.05, 
        direction='nearest'
    )

    # Identificar valores que não foram encontrados na tabela auxiliar interna
    nao_encontrados = merged[merged['Valor final'].isna()]
    
    if not nao_encontrados.empty:
        print("AVISO: Os seguintes valores de venda sao novos e nao estao mapeados no codigo")
        print("       (mesmo considerando 5 centavos de tolerancia):")
        valores_unicos = nao_encontrados['Valor'].unique()
        for v in valores_unicos:
            qtd = len(nao_encontrados[nao_encontrados['Valor'] == v])
            print(f"  - R$ {v:.2f} (aparece {qtd} vez(es))")
        print("-> Para adicionar esses valores, voce precisara atualizar a lista 'dados_auxiliares' dentro do arquivo Python.\n")

    # Filtrar apenas os que encontraram correspondência para calcular o lucro
    encontrados = merged.dropna(subset=['Valor final']).copy()
    
    if encontrados.empty:
        print("Nenhuma venda correspondente encontrada. O arquivo não tem vendas ou os valores são todos novos.")
        return

    # Criar uma coluna para a semana (começando na segunda-feira)
    encontrados['Semana'] = encontrados['Data'].dt.to_period('W-MON').apply(lambda r: r.start_time)
    
    # Agrupar por semana e calcular os totais
    resumo_semanal = encontrados.groupby('Semana').agg(
        Qtd_Vendas=('Valor', 'count'),
        Total_Vendido=('Valor', 'sum'),
        Custo_Total=('Custo', 'sum'),
        Lucro_Total=('Lucro', 'sum')
    ).reset_index()

    print("RESUMO DE LUCRO SEMANAL:")
    print("-" * 60)
    for index, row in resumo_semanal.iterrows():
        inicio_semana = row['Semana'].strftime('%d/%m/%Y')
        fim_semana = (row['Semana'] + pd.Timedelta(days=6)).strftime('%d/%m/%Y')
        
        print(f"Semana de {inicio_semana} ate {fim_semana}:")
        print(f"  Vendas mapeadas com sucesso:   {int(row['Qtd_Vendas'])}")
        print(f"  Total Vendido (bruto):         R$ {row['Total_Vendido']:.2f}")
        print(f"  Custo Total:                   R$ {row['Custo_Total']:.2f}")
        print(f"  Lucro Total da Semana:         R$ {row['Lucro_Total']:.2f}")
        print("-" * 60)
        
    lucro_total_geral = resumo_semanal['Lucro_Total'].sum()
    print(f"\nLUCRO TOTAL DO PERIODO (apenas vendas mapeadas): R$ {lucro_total_geral:.2f}")

    # ==========================================
    # GERAÇÃO DA PLANILHA EXCEL DE FECHAMENTO
    # ==========================================
    nome_arquivo_base = os.path.basename(caminho_arquivo)
    nome_sem_extensao = os.path.splitext(nome_arquivo_base)[0]
    
    # Usa o nome pedido se for o arquivo 11_18, senão cria um nome genérico
    if "11_18" in nome_sem_extensao:
        caminho_saida = "Fechamento Semanal Mai 11 a 18.xlsx"
    else:
        caminho_saida = f"Fechamento_{nome_sem_extensao}.xlsx"
        
    caminho_saida_completo = os.path.join(os.path.dirname(os.path.abspath(caminho_arquivo)), caminho_saida)
    
    # 1. Preparar Aba de Resumo
    df_resumo_export = resumo_semanal.copy()
    df_resumo_export['Semana_Inicio'] = df_resumo_export['Semana'].dt.strftime('%d/%m/%Y')
    df_resumo_export['Semana_Fim'] = (df_resumo_export['Semana'] + pd.Timedelta(days=6)).dt.strftime('%d/%m/%Y')
    df_resumo_export['Periodo'] = df_resumo_export['Semana_Inicio'] + " ate " + df_resumo_export['Semana_Fim']
    cols_resumo = ['Periodo', 'Qtd_Vendas', 'Total_Vendido', 'Custo_Total', 'Lucro_Total']
    df_resumo_export = df_resumo_export[cols_resumo]
    
    linha_total = pd.DataFrame([{
        'Periodo': 'TOTAL GERAL',
        'Qtd_Vendas': df_resumo_export['Qtd_Vendas'].sum(),
        'Total_Vendido': df_resumo_export['Total_Vendido'].sum(),
        'Custo_Total': df_resumo_export['Custo_Total'].sum(),
        'Lucro_Total': df_resumo_export['Lucro_Total'].sum()
    }])
    df_resumo_export = pd.concat([df_resumo_export, linha_total], ignore_index=True)

    # 2. Preparar Aba de Todos os Pedidos
    df_pedidos_export = encontrados.copy()
    df_pedidos_export['Data'] = df_pedidos_export['Data'].dt.strftime('%d/%m/%Y %H:%M:%S')
    df_pedidos_export = df_pedidos_export[['Data', 'Valor', 'Custo', 'Lucro']].copy()
    df_pedidos_export.rename(columns={'Valor': 'Valor Venda'}, inplace=True)
    
    # 3. Exportar com formatação premium e estilizada
    try:
        # Definir estilos visuais de alta qualidade
        font_cabecalho = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        font_dados = Font(name="Segoe UI", size=10)
        font_total = Font(name="Segoe UI", size=10, bold=True)
        
        # Paleta de Cores: Azul Marinho Clássico para Finanças
        fill_cabecalho = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid") # Navy Blue
        fill_zebra = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid") # Azul acinzentado ultra suave
        fill_total = PatternFill(start_color="E1EBF5", end_color="E1EBF5", fill_type="solid") # Azul suave de realce
        
        align_centro = Alignment(horizontal="center", vertical="center")
        align_esquerda = Alignment(horizontal="left", vertical="center")
        align_direita = Alignment(horizontal="right", vertical="center")
        
        borda_sutil = Side(style="thin", color="E0E0E0")
        borda_dados = Border(left=borda_sutil, right=borda_sutil, top=borda_sutil, bottom=borda_sutil)
        
        borda_total_top = Side(style="thin", color="1B365D")
        borda_total_bottom = Side(style="double", color="1B365D")
        borda_total = Border(left=borda_sutil, right=borda_sutil, top=borda_total_top, bottom=borda_total_bottom)
        
        with pd.ExcelWriter(caminho_saida_completo, engine='openpyxl') as writer:
            # Exportar os DataFrames base primeiro
            df_resumo_export.to_excel(writer, sheet_name='Resumo Semanal', index=False)
            df_pedidos_export.to_excel(writer, sheet_name='Todos os Pedidos', index=False)
            if not nao_encontrados.empty:
                df_nao_encontrados = pd.DataFrame({
                    'Valor da Venda (Nao Mapeado)': nao_encontrados['Valor'].value_counts().index,
                    'Quantidade de Vezes': nao_encontrados['Valor'].value_counts().values
                })
                df_nao_encontrados.to_excel(writer, sheet_name='Valores Pendentes', index=False)
                
            workbook = writer.book
            
            # ----------------------------------------------------
            # 3.1. Estilizar Aba 'Resumo Semanal'
            # ----------------------------------------------------
            ws_resumo = writer.sheets['Resumo Semanal']
            ws_resumo.views.sheetView[0].showGridLines = True
            
            # Cabeçalho
            ws_resumo.row_dimensions[1].height = 26
            for col in range(1, 6):
                cell = ws_resumo.cell(row=1, column=col)
                cell.font = font_cabecalho
                cell.fill = fill_cabecalho
                cell.alignment = align_centro
                cell.border = Border(left=Side(style="thin", color="2C4D75"), right=Side(style="thin", color="2C4D75"))
                
            # Dados
            max_rows_resumo = ws_resumo.max_row
            for row in range(2, max_rows_resumo + 1):
                ws_resumo.row_dimensions[row].height = 20
                eh_total = (row == max_rows_resumo)
                
                for col in range(1, 6):
                    cell = ws_resumo.cell(row=row, column=col)
                    
                    if eh_total:
                        cell.font = font_total
                        cell.fill = fill_total
                        cell.border = borda_total
                    else:
                        cell.font = font_dados
                        cell.border = borda_dados
                        if row % 2 == 1:
                            cell.fill = fill_zebra
                            
                    # Alinhamento e Formato de Número
                    if col == 1:
                        cell.alignment = align_centro
                    elif col == 2:
                        cell.alignment = align_centro
                        cell.number_format = '#,##0'
                    else:
                        cell.alignment = align_direita
                        cell.number_format = 'R$ #,##0.00'
                        
            # Ajuste de largura das colunas
            for col in ws_resumo.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws_resumo.column_dimensions[col_letter].width = max(max_len + 5, 13)
                
            # ----------------------------------------------------
            # 3.2. Estilizar Aba 'Todos os Pedidos'
            # ----------------------------------------------------
            ws_pedidos = writer.sheets['Todos os Pedidos']
            ws_pedidos.views.sheetView[0].showGridLines = True
            ws_pedidos.row_dimensions[1].height = 26
            
            # Cabeçalho
            for col in range(1, 5):
                cell = ws_pedidos.cell(row=1, column=col)
                cell.font = font_cabecalho
                cell.fill = fill_cabecalho
                cell.alignment = align_centro
                cell.border = Border(left=Side(style="thin", color="2C4D75"), right=Side(style="thin", color="2C4D75"))
                
            # Dados
            max_rows_pedidos = ws_pedidos.max_row
            for row in range(2, max_rows_pedidos + 1):
                ws_pedidos.row_dimensions[row].height = 19
                for col in range(1, 5):
                    cell = ws_pedidos.cell(row=row, column=col)
                    cell.font = font_dados
                    cell.border = borda_dados
                    if row % 2 == 1:
                        cell.fill = fill_zebra
                        
                    if col == 1:
                        cell.alignment = align_centro
                    else:
                        cell.alignment = align_direita
                        cell.number_format = 'R$ #,##0.00'
                        
            # Ajuste de largura das colunas
            for col in ws_pedidos.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws_pedidos.column_dimensions[col_letter].width = max(max_len + 5, 14)
                
            # ----------------------------------------------------
            # 3.3. Estilizar Aba 'Valores Pendentes'
            # ----------------------------------------------------
            if not nao_encontrados.empty:
                ws_pendentes = writer.sheets['Valores Pendentes']
                ws_pendentes.views.sheetView[0].showGridLines = True
                ws_pendentes.row_dimensions[1].height = 26
                
                # Cabeçalho em Vermelho Sutil
                fill_vermelho = PatternFill(start_color="9C0006", end_color="9C0006", fill_type="solid")
                for col in range(1, 3):
                    cell = ws_pendentes.cell(row=1, column=col)
                    cell.font = font_cabecalho
                    cell.fill = fill_vermelho
                    cell.alignment = align_centro
                    cell.border = Border(left=Side(style="thin", color="B22222"), right=Side(style="thin", color="B22222"))
                    
                # Dados
                max_rows_pendentes = ws_pendentes.max_row
                for row in range(2, max_rows_pendentes + 1):
                    ws_pendentes.row_dimensions[row].height = 19
                    for col in range(1, 3):
                        cell = ws_pendentes.cell(row=row, column=col)
                        cell.font = font_dados
                        cell.border = borda_dados
                        if row % 2 == 1:
                            cell.fill = fill_zebra
                            
                        if col == 1:
                            cell.alignment = align_direita
                            cell.number_format = 'R$ #,##0.00'
                        else:
                            cell.alignment = align_centro
                            cell.number_format = '#,##0'
                            
                # Ajuste de largura das colunas
                for col in ws_pendentes.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    col_letter = get_column_letter(col[0].column)
                    ws_pendentes.column_dimensions[col_letter].width = max(max_len + 6, 16)
                    
        print(f"\nPLANILHA GERADA E FORMATADA: {caminho_saida}")
        print("-> O Excel com tudo descrito foi estilizado e salvo com sucesso!")
    except Exception as e:
        print(f"\nErro ao gerar a planilha Excel: {e}")

if __name__ == "__main__":
    if len(sys.argv) > 1:
        caminho = sys.argv[1]
        if os.path.exists(caminho):
            calcular_lucro_semanal(caminho)
        else:
            print(f"Arquivo não encontrado: {caminho}")
            print("Verifique se o caminho ou o nome do arquivo está correto.")
    else:
        print("Por favor, informe o nome ou o caminho do arquivo.")
        print("Exemplo: python calculadora_lucro.py 11_18.xlsx")
